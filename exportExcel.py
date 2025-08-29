from typing import List
from io import BytesIO
from openpyxl import load_workbook
from fastapi.responses import StreamingResponse
from models import StudentInterface

def export_data(list_of_students: List[StudentInterface]):
    file_path = "TEMPLATE.xlsx"
    wb = load_workbook(file_path)
    ws = wb.active

    headers = [
        "STUDENT ID",
        "LASTNAME",
        "FIRSTNAME",
        "FAMILY INCOME",
        "TYPE OF SENIOR HIGH SCHOOL",
        "PROGRAM ENROLLED",
        "MUNICIPALITY OF ORIGIN",
        "GRADE 12 GWA",
        "SEX"
    ]

    # Write headers starting in row 7
    row_num = 7
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=row_num, column=col_num, value=header)

    # Insert student data starting in row 8
    for idx, student in enumerate(list_of_students, start=8):
        ws.cell(row=idx, column=1, value=student.ID)
        ws.cell(row=idx, column=2, value=student.Lastname)
        ws.cell(row=idx, column=3, value=student.Firstname)
        ws.cell(row=idx, column=4, value=student.FamilyIncome)
        ws.cell(row=idx, column=5, value=student.TypeofSeniorHighSchool)
        ws.cell(row=idx, column=6, value=student.ProgramEnrolled)
        ws.cell(row=idx, column=7, value=student.MunicipalityOfOrigin)
        ws.cell(row=idx, column=8, value=student.Grade12GWA)
        ws.cell(row=idx, column=9, value=student.Sex)

    # Save workbook to in-memory bytes buffer
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Return StreamingResponse directly
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="updated_file.xlsx"'}
    )
